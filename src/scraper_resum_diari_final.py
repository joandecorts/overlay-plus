#!/usr/bin/env python3
# scraper_resum_diari_final.py - Genera Excel formatat, CSV i JSON
import requests
from bs4 import BeautifulSoup
import pandas as pd
import sys
import time
from datetime import datetime
import json
from pathlib import Path

import sys
from pathlib import Path

# 🔧 SOLUCIÓ: Afegir la carpeta 'config' al camí de cerca de Python
sys.path.insert(0, str(Path(__file__).parent.parent / 'config'))

# --- IMPORTACIÓ DE LA CONFIGURACIÓ CENTRAL ---
try:
    from config_banner import STATIONS, TODAY, DATA_DIR
    print("✅ Configuració importada correctament des de 'config_banner.py'")
except ImportError as e:
    print(f"❌ Error important la configuració: {e}")
    sys.exit(1)

# --- CONFIGURACIÓ ---
DIA_CONSULTA = TODAY.strftime("%Y-%m-%d")
BASE_URL = "https://www.meteo.cat/observacions/xema/dades"
DELAI_ENTRE_PETICIONS = 1
HORA_CONSULTA = "T09:00Z"  # Hora per defecte que funciona

# Diccionari de variables (9 variables - sense DIRECCIO_VENT)
MAP_VARIABLES = {
    'Temperatura mitjana': 'TEMPERATURA_MITJANA_DIA',
    'Temperatura màxima': 'TEMPERATURA_MAXIMA_DIA',
    'Temperatura mínima': 'TEMPERATURA_MINIMA_DIA',
    'Humitat relativa mitjana': 'HUMITAT_MITJANA_DIA',
    'Precipitació acumulada': 'PRECIPITACIO_ACUM_DIA',
    'Gruix de neu màxim': 'GRUIX_NEU_MAX',
    'Ratxa màxima del vent': 'RATXA_VENT_MAX',
    'Irradiació solar global': 'RADIACIO_GLOBAL',
    'Pressió atmosfèrica': 'PRESSIO_ATMOSFERICA'  # ✅ MANTINGUT
    # ❌ ELIMINAT: 'Direcció del vent': 'DIRECCIO_VENT'
}

def obtenir_info_estacio(codi_estacio):
    """Obtenir nom de l'estació des de config_banner.py"""
    for estacio in STATIONS:
        if estacio.get('code') == codi_estacio:
            return {
                'nom': estacio.get('display_name', estacio.get('name', codi_estacio)),
                'nom_original': estacio.get('name', '')
            }
    return {'nom': codi_estacio, 'nom_original': ''}

def neteja_valor(text):
    """Netega i formata el text"""
    if not text or text in ['(s/d)', '-', '', 'N/D', 's/d']:
        return ''
    text = text.replace('MJ/m 2', 'MJ/m2')
    return ' '.join(text.split())

def extreu_resum_diari_per_estacio(codi_estacio, dia):
    """Extreu dades d'una estació"""
    # URL amb hora configurable
    url = f"{BASE_URL}?codi={codi_estacio}&dia={dia}{HORA_CONSULTA}"
    
    # Obtenir info de l'estació
    info_estacio = obtenir_info_estacio(codi_estacio)
    
    # Inicialitzar amb totes les columnes
    resultats = {nom_var: '' for nom_var in MAP_VARIABLES.values()}
    resultats['ID_ESTAC'] = codi_estacio
    resultats['NOM_ESTACIO'] = info_estacio['nom']
    resultats['NOM_ORIGINAL'] = info_estacio['nom_original']
    resultats['DATA_DIA'] = dia
    resultats['DATA_EXTRACCIO'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    resultats['URL_FONT'] = url

    try:
        resposta = requests.get(url, timeout=15)
        resposta.raise_for_status()
    except:
        return resultats

    soup = BeautifulSoup(resposta.text, 'html.parser')
    taules = soup.find_all('table')
    taula_trobada = None

    # Buscar taula
    for taula in taules:
        text_anterior = taula.find_previous(['h2', 'h3', 'strong', 'b'])
        if text_anterior and 'Resum diari' in str(text_anterior):
            taula_trobada = taula
            break
    
    if not taula_trobada:
        for taula in taules:
            if 'Temperatura mitjana' in str(taula):
                taula_trobada = taula
                break

    if not taula_trobada:
        return resultats

    # Processar files
    files = taula_trobada.find_all('tr')
    for fila in files:
        cel·les = fila.find_all(['td', 'th'])
        if len(cel·les) >= 2:
            text_clau = cel·les[0].get_text(strip=True)
            for text_web, nom_variable in MAP_VARIABLES.items():
                if text_web in text_clau:
                    valor = ' '.join(cel.get_text(strip=True, separator=' ') for cel in cel·les[1:])
                    resultats[nom_variable] = neteja_valor(valor)
                    break

    return resultats

def executa_scraping_estacions(llista_estacions, dia):
    """Executa per totes les estacions"""
    totes_dades = []
    
    print(f"\n🚀 Iniciant scraping per a {len(llista_estacions)} estacions...")
    print(f"📅 Data: {dia}{HORA_CONSULTA}")
    print("-" * 70)

    for idx, estacio in enumerate(llista_estacions, 1):
        codi = estacio.get('code')
        nom = estacio.get('display_name', estacio.get('name', codi))
        print(f"[{idx:3}/{len(llista_estacions)}] 🔍 {nom} ({codi})...", end=' ', flush=True)

        dades = extreu_resum_diari_per_estacio(codi, dia)
        totes_dades.append(dades)
        
        # ✅ COMPTATGE CORREGIT: comptem totes les variables de MAP_VARIABLES que tinguin valor
        dades_trobades = sum(1 for nom_var in MAP_VARIABLES.values() if dades.get(nom_var))
        print(f"{dades_trobades} vars" if dades_trobades > 0 else "sense dades")

        time.sleep(DELAI_ENTRE_PETICIONS)

    return totes_dades

def genera_excel_formatat(df, ruta_excel):
    """Genera un arxiu Excel amb format professional"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        writer = pd.ExcelWriter(ruta_excel, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Dades_Meteo')
        worksheet = writer.sheets['Dades_Meteo']
        
        # Estils
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar format als capçaleres
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # Ajustar amplada de columna
            col_letter = get_column_letter(col)
            max_length = max(
                df.iloc[:, col-1].astype(str).map(len).max(),
                len(str(df.columns[col-1]))
            )
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 30)
        
        # Aplicar borders a totes les cel·les
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, max_col=len(df.columns)):
            for cell in row:
                cell.border = border
                if cell.column in [1, 2, 4]:  # ID, NOM, DATA
                    cell.alignment = Alignment(vertical='center')
        
        # Congelar panells (ID i NOM visibles)
        worksheet.freeze_panes = 'C2'
        
        writer.close()
        print(f"💼 Excel formatat guardat: {ruta_excel}")
        return True
        
    except ImportError:
        print("⚠️  openpyxl no instal·lat, generant Excel sense format")
        df.to_excel(ruta_excel, index=False)
        return True
    except Exception as e:
        print(f"❌ Error generant Excel: {e}")
        return False

def guarda_tots_formats(totes_dades, data_consulta):
    """Guarda en tots els formats: Excel, CSV i JSON"""
    if not totes_dades:
        return None, None, None

    # Crear DataFrame
    df = pd.DataFrame(totes_dades)
    
    # ORDENAR COLUMNES: Metadades primer, després variables meteorològiques
    columnes_metadades = [
        'ID_ESTAC', 'NOM_ESTACIO', 'NOM_ORIGINAL',
        'DATA_DIA', 'DATA_EXTRACCIO', 'URL_FONT'
    ]
    
    # Separar variables meteorològiques (només les que estan a MAP_VARIABLES)
    columnes_meteo = [c for c in df.columns if c in MAP_VARIABLES.values()]
    
    # Ordenar variables meteorològiques alfabèticament
    columnes_meteo_ordenades = sorted(columnes_meteo)
    
    # Columnes finals
    columnes_ordenades = columnes_metadades + columnes_meteo_ordenades
    df = df[columnes_ordenades]
    
    # Crear directori si no existeix
    directori_dades = Path(DATA_DIR)
    directori_dades.mkdir(parents=True, exist_ok=True)
    
    # Timestamp per als noms
    timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
    nom_base = f"dades_dia_{timestamp}"
    
    # 1. EXCEL FORMATAT (.xlsx)
    nom_excel = f"{nom_base}.xlsx"
    ruta_excel = directori_dades / nom_excel
    excel_ok = genera_excel_formatat(df, ruta_excel)
    
    # 2. CSV SIMPLE (.csv)
    nom_csv = f"{nom_base}.csv"
    ruta_csv = directori_dades / nom_csv
    df.to_csv(ruta_csv, index=False, encoding='utf-8-sig')
    print(f"📄 CSV simple guardat: {ruta_csv}")
    
    # 3. JSON PER AL BANNER
    nom_json = f"{nom_base}.json"
    ruta_json = directori_dades / nom_json
    
    dades_json = {
        'metadata': {
            'data_consulta': data_consulta,
            'hora_consulta': HORA_CONSULTA,
            'data_extractcio': datetime.now().isoformat(),
            'total_estacions': len(totes_dades),
            'formats_generats': ['Excel', 'CSV', 'JSON'],
            'variables_capturades': list(MAP_VARIABLES.values())
        },
        'estacions': totes_dades
    }
    
    with open(ruta_json, 'w', encoding='utf-8') as f:
        json.dump(dades_json, f, ensure_ascii=False, indent=2)
    print(f"📋 JSON per al banner: {ruta_json}")
    
    return ruta_excel, ruta_csv, ruta_json if excel_ok else (None, ruta_csv, ruta_json)

# --- EXECUCIÓ PRINCIPAL ---
if __name__ == "__main__":
    print("\n" + "="*70)
    print("🌤️  SCRAPER DE RESUM DIARI - Excel, CSV i JSON")
    print("="*70)
    print(f"⏰ Hora utilitzada: {HORA_CONSULTA}")

    # OPCIONS
    print(f"📋 Estacions disponibles: {len(STATIONS)}")
    
    print("\n🎯 MODES D'EXECUCIÓ:")
    print("1. TOTES les estacions")
    print("2. Mode PROVES (Z3, XI, XJ, C6)")
    print("3. Excloure UO i problemàtiques")
    
    try:
        opcio = int(input("\n👉 Selecciona opció (1-3): ").strip() or "1")
    except:
        opcio = 1
    
    if opcio == 1:
        estacions_a_processar = STATIONS
    elif opcio == 2:
        codis_prova = ['Z3', 'XI', 'XJ', 'C6']
        estacions_a_processar = [s for s in STATIONS if s['code'] in codis_prova]
    elif opcio == 3:
        codis_excloure = ['UO']  # Afegeix més aquí si cal
        estacions_a_processar = [s for s in STATIONS if s['code'] not in codis_excloure]
    else:
        estacions_a_processar = STATIONS
    
    print(f"\n▶️  Estacions seleccionades: {len(estacions_a_processar)}")
    
    # Preguntar per canviar l'hora si es vol
    canviar_hora = input(f"▶️  Canviar hora de consulta ({HORA_CONSULTA})? (s/n): ").strip().lower()
    if canviar_hora == 's':
        nova_hora = input("Nova hora (ex: T10:00Z): ").strip()
        if nova_hora:
            HORA_CONSULTA = nova_hora
    
    continuar = input("▶️  Continuar amb l'execució? (s/n): ").strip().lower()
    if continuar != 's':
        print("⏹️  Execució cancel·lada.")
        sys.exit(0)

    # SCRAPING
    dades = executa_scraping_estacions(estacions_a_processar, DIA_CONSULTA)

    # GUARDAR I INFORME
    if dades:
        ruta_excel, ruta_csv, ruta_json = guarda_tots_formats(dades, DIA_CONSULTA)
        
        print("\n" + "="*70)
        print("📊 RESULTATS GENERATS")
        print("="*70)
        
        # Estadístiques
        # Comptem quantes de les variables de MAP_VARIABLES tenen dades en alguna estació
        df_temp = pd.DataFrame(dades)
        vars_amb_dades = sum(1 for var in MAP_VARIABLES.values() 
                            if var in df_temp.columns and df_temp[var].notna().any())
        
        print(f"✅ Estacions processades: {len(dades)}")
        print(f"📈 Variables amb dades: {vars_amb_dades}/{len(MAP_VARIABLES)}")
        print(f"📁 Directori: {DATA_DIR}")
        print()
        print(f"📊 1. EXCEL formatat:   {Path(ruta_excel).name if ruta_excel else 'No generat'}")
        print(f"📄 2. CSV simple:       {Path(ruta_csv).name}")
        print(f"📋 3. JSON per banner:  {Path(ruta_json).name}")
        print()
        print("💡 Obre l'Excel per veure el format professional!")
        print("="*70)
        
        # Mostrar mostra ràpida
        print("\n👁️  MOSTRA RÀPIDA (primera estació):")
        print("-" * 50)
        if dades:
            primera = dades[0]
            for clau in ['ID_ESTAC', 'NOM_ESTACIO', 'DATA_DIA', 
                        'TEMPERATURA_MITJANA_DIA', 'PRECIPITACIO_ACUM_DIA', 'URL_FONT']:
                if clau in primera:
                    valor = primera[clau] if primera[clau] else '(buit)'
                    print(f"{clau:<25}: {valor}")
        
    else:

        print("\n❌ No s'han obtingut dades. Revisa la connexió.")
